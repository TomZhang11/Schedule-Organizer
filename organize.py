# color codes
# red background FFCCCB
# banana background FFFFE0
# blue background ADD8E6
# green background 90EE90
# red text FF0000
# purple text A020F0
# pink text FF44FC
# orange text E86C0C

from glob import glob
from os.path import exists, basename
from os import makedirs
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from re import sub

def missing_files():
    error=False
    class_files=glob("Input/*classes*.txt")
    if not class_files:
        print("error: missing the classes file")
        error=True
    elif len(class_files)>1:
        print("error: there are more than 1 class files")
        error=True
    if not exists("Input/departments.txt"):
        print("error: missing the departments file")
        error=True
    if not exists("Input/heights.txt"):
        print("error: missing the heights file")
        error=True
    if error:
        print("program terminated, no changes have been made")
        input("press Enter to exit")
    return error, basename(class_files[0])[:-4]

def read_classes_file(classes_filename):
    classes=[]
    with open("Input/"+classes_filename+".txt", "r", encoding="utf-8") as file:
        next(file)
        for line in file:
            if not line.strip():
                continue
            class_details=line.strip().split(",")
            if len(class_details)<3:
                print(f"warning: line in wrong format in the classes file:")
                print(f"  {line}", end="")
                print("  program skipped this line and continued to run")
                continue
            class_info={
                "class_code":class_details[0].strip('"'),
                "semester":class_details[1].strip('"'),
                "period":class_details[2].strip('"')
            }
            classes.append(class_info)
    return classes

def read_departments_file():
    departments=[]
    department_values={}
    with open("Input/departments.txt", "r", encoding="utf-8") as file:
        line_number=0
        for line in file:
            if not line.strip():
                continue
            elif ":" in line:
                department, class_codes=line.split(":", 1)
                departments.append(department.strip())
                for class_code in class_codes.split():
                    department_values[class_code]=line_number
                line_number+=1
            else:
                print(f"warning: line in wrong format in departments.txt")
                print(f"  {line}")
                print("  program skipped this line and continued to run")
    departments.append("Totals")
    return departments, department_values

def get_heights():
    period_height=5
    lunch_height=1
    stage=0
    with open("Input/heights.txt", "r", encoding="utf-8") as file:
        for line in file:
            if ":" in line:
                height=line.split(":")[1].strip()
                if height.isdigit():
                    if stage==0:
                        period_height=int(height)
                    else:
                        lunch_height=int(height)
                        return period_height+1, lunch_height+1
                else:
                    print("warning: invalid height in heights.txt")
                    print(f"  {height}")
                    print("  program used the default height")
                stage+=1
    return period_height+1, lunch_height+1

def isfloat(s):
    try:
        float(s)
        return True
    except:
        return False

def read_values_file():
    class_values={}
    if exists("Input/class values.xlsx"):
        values_file=load_workbook("Input/class values.xlsx")
        values_sheet=values_file.active
        row=2
        while not values_sheet["A"+str(row)].value is None:
            class_code=values_sheet["A"+str(row)].value
            values=[]
            max_val=0
            max_grade=0
            valid=True
            for col in range(2, 6):
                value=values_sheet.cell(row=row, column=col).value
                if value is None:
                    values.append(0)
                elif isfloat(value):
                    if value>max_val:
                        max_val=value
                        max_grade=col-2
                    elif value==max_val:
                        max_grade=-1
                    values.append(value)
                else:
                    print(f"warning: invalid entry {value} in the class values file")
                    print("  program ignored this row and continued to run")
                    valid=False
            if valid:
                restrictions=values_sheet["F"+str(row)].value
                if restrictions==None:
                    restrictions=""
                class_values[class_code]=(values, str(restrictions), max_grade)
                row+=1
    else:
        print("warning: the class values file does not exist")
        print("  the program used default values for all classes")
    return class_values

def setup_schedule(schedule, departments, period_height, lunch_height):
    max_col=len(departments)*5
    max_row=period_height*10+lunch_height*2

    # set column widths
    for col in range(1, max_col+1):
        schedule.column_dimensions[get_column_letter(col)].width=7.33

    # write department headers
    for i, department in enumerate(departments):
        left_col=i*5+2
        right_col=i*5+5
        schedule.merge_cells(start_row=1, start_column=left_col, end_row=1, end_column=right_col)
        cell=schedule.cell(row=1, column=left_col)
        cell.value=department
        cell.font=Font(size=16)
        cell.alignment=Alignment(horizontal='center')

    # write periods
    periods=["P1", "P2", "P3", "Lunch", "P4", "P5"] * 2
    row_numbers=[2] # [2, 8, 14, 20, 22, 28, 34, 40, 46, 52, 54, 60]
    row=2
    for i in range(11):
        if i==3 or i==9:
            row+=lunch_height
        else:
            row+=period_height
        row_numbers.append(row)
    for period, row in zip(periods, row_numbers):
        schedule["A"+str(row)]=period

    # background colors
    for col in range(2, max_col-2, 5):
        for row in range(2, max_row+1):
            if row+1 in row_numbers:
                continue
            schedule.cell(row=row, column=col).fill=PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
            schedule.cell(row=row, column=col+1).fill=PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            schedule.cell(row=row, column=col+2).fill=PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            schedule.cell(row=row, column=col+3).fill=PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    # grey borders between departments
    for i in range(6, max_col-3, 5):
        col=get_column_letter(i)
        schedule.column_dimensions[col].width=2
        for row in range(1, max_row+1):
            schedule[col+str(row)].fill=PatternFill(start_color="5A5A5A", end_color="5A5A5A", fill_type="solid")

    # grey borders between periods
    for i in range(1, 12):
        row=row_numbers[i]-1
        schedule.row_dimensions[row].height=10
        for col in range(1, max_col+1):
            schedule.cell(row=row, column=col).fill=PatternFill(start_color="5A5A5A", end_color="5A5A5A", fill_type="solid")
    semester_border=row_numbers[6]-1
    schedule.row_dimensions[semester_border].height=20

    # color indicators for grades
    colors=["FFCCCB", "FFFFE0", "ADD8E6", "90EE90"]
    for col, color in enumerate(colors):
        schedule.cell(row=max_row+2, column=col+2).fill=PatternFill(start_color=color, end_color=color, fill_type="solid")
        schedule.cell(row=max_row+3, column=col+2).value=f"Gr.{9+col}"
    return row_numbers

def clean_code(code):
    # using a regular expression to remove the grp part at the end
    code = sub(r'-grp\s*\d*$', '', code)
    # using a regular expression to remove three digit codes at the end
    code = sub(r'-\d{3}$', '', code)
    return code

def enter_class(schedule, class_code, col, row, semester, period, day, row_numbers):
    # finding an empty row
    while not row+1 in row_numbers:
        cell=schedule.cell(row=row, column=col)
        if cell.value is None:
            cell.value=clean_code(class_code)
            # purple color for group classes
            if "grp" in class_code:
                cell.font=Font(color="A020F0")
                if semester=="FY":
                    cell.value=class_code.split("grp")[0]+"grp"
            if semester=="FY":
                # pink color for D1 classes
                if day=="D1":
                    cell.font=Font(color="FF44FC")
                # orange color for D2 classes
                elif day=="D2":
                    cell.font=Font(color="E86C0C")
                # red color for D1-D2 classes
                else:
                    cell.font=Font(color="FF0000")
            return
        row+=1
    print("warning: number of classes exceeded period height, the following class could not be entered into the sheet:")
    print(f"  {class_code}  semester: {semester}  period: {period}")

def update_totals(totals, class_values, class_code, index, grade_value, semester, group_classes, class_period):
    cleaned_code=clean_code(class_code)
    if cleaned_code in class_values and semester in class_values[cleaned_code][1]:
        for i in range(4):
            totals[index][i]+=class_values[cleaned_code][0][i]
    elif "grp" in class_code:
        if "grp" in class_code:
            code=class_code.split("grp")[1]
            if code in group_classes:
                group_classes[code][0][grade_value]+=1
                group_classes[code][3]+=1
            else:
                group_classes[code]=[[0, 0, 0, 0], semester, class_period, 1]
                group_classes[code][0][grade_value]+=1
    else:
        totals[index][grade_value]+=1

def populate_sheet(schedule, classes, departments, department_values, class_values, row_numbers):
    french_department_exists="FRIMM" in departments
    semester1_rows={"1":row_numbers[0], "2":row_numbers[1], "3":row_numbers[2], "L":row_numbers[3], "4":row_numbers[4], "5":row_numbers[5]}
    semester2_rows={"1":row_numbers[6], "2":row_numbers[7], "3":row_numbers[8], "L":row_numbers[9], "4":row_numbers[10], "5":row_numbers[11]}
    totals=[]
    group_classes={}
    for _ in range(12):
        totals.append([0, 0, 0, 0])
    period_indexes={"1":0, "2":1, "3":2, "L":3, "4":4, "5":5}
    for i in classes:
        # determining the correct column for the class
        three_letter_code=i["class_code"][:3]
        cleaned_code=clean_code(i["class_code"])
        grade_str=i["class_code"][3:5]
        grade_value=0
        if grade_str.isdigit():
            grade_value=int(grade_str)//10
        if cleaned_code in class_values:
            value=class_values[cleaned_code][2]
            if value!=-1:
                grade_value=value
        if i["class_code"][5]=="F":
            if not french_department_exists:
                print(f"warning: department FRIMM is needed for the class {i['class_code']}")
                print("  program ignored this class and continued to run")
                continue
            col=departments.index("FRIMM")*5+2+grade_value
        elif three_letter_code in department_values:
            col=department_values[three_letter_code]*5+2+grade_value
        else:
            print(f"warning: the class {i['class_code']} is not assigned to a department")
            print("  program ignored this class and continued to run")
            continue
        # handling semester and period format errors
        class_period=i["period"][0]
        if not i["semester"] in ["S1", "S2", "FY"]:
            print(f"warning: invalid semester {i['semester']} for the class {i['class_code']}")
            print("  program ignored this class and continued to run")
            continue
        elif not class_period in "12345L":
            print(f"warning: invalid period {i['period']} for the class {i['class_code']}")
            print("  program ignored this class and continued to run")
            continue
        # determining the correct row for the class and entering the class onto the excel sheet
        day=i["period"].split("(")[1][:-1]
        if i["semester"] in ["S1", "FY"]:
            row=semester1_rows[class_period]
            enter_class(schedule, i["class_code"], col, row, i["semester"], class_period, day, row_numbers)
            update_totals(totals, class_values, i["class_code"], period_indexes[class_period], grade_value, i["semester"], group_classes, class_period)
        if i["semester"] in ["S2", "FY"]:
            row=semester2_rows[class_period]
            enter_class(schedule, i["class_code"], col, row, i["semester"], class_period, day, row_numbers)
            update_totals(totals, class_values, i["class_code"], period_indexes[class_period]+6, grade_value, i["semester"], group_classes, class_period)
        if "grp" in i["class_code"]:
            code=i["class_code"].split("grp")[1]
            if code in group_classes:
                group_classes[code][0][grade_value]+=1
                group_classes[code][3]+=1
            else:
                group_classes[code]=[[0, 0, 0, 0], i["semester"], class_period, 1]
                group_classes[code][0][grade_value]+=1
    for code in group_classes:
        index=period_indexes[group_classes[code][2]]
        if group_classes[code][1] in ["S1", "FY"]:
            for i in range(4):
                totals[index][i]+=group_classes[code][0][i]/group_classes[code][3]
        if group_classes[code][1] in ["S2", "FY"]:
            for i in range(4):
                totals[index+6][i]+=group_classes[code][0][i]/group_classes[code][3]
    for i in range(12):
        for j in range(4):
            schedule.cell(row=row_numbers[i], column=len(departments)*5-3+j).value=totals[i][j]

def save(excel_file, classes_filename):
    # date_time=datetime.now().strftime("%b %d %H:%M:%S")
    if not exists("Output"):
        makedirs("Output")
    excel_file.save(f"Output/{classes_filename.replace('classes', 'schedule')}.xlsx")
    input("program ran sucessfully, press Enter to exit")

def main():
    error, classes_filename=missing_files()
    if error:
        return
    classes=read_classes_file(classes_filename)
    departments, department_values=read_departments_file()
    period_height, lunch_height=get_heights()
    class_values=read_values_file()
    excel_file=Workbook()
    schedule=excel_file.active
    row_numbers=setup_schedule(schedule, departments, period_height, lunch_height)
    populate_sheet(schedule, classes, departments, department_values, class_values , row_numbers)
    save(excel_file, classes_filename)

main()